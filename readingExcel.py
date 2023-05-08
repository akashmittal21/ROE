import openpyxl
import random
import datetime

"""
Plan Selection Basis:
-> If the spouse or children have no insurance:
    - If the employee has no spouse or children, then all the plans selected need to be single plans
    - If the employee has a spouse and no children or no spouse and one child, then all the plans need to be couple plans
    - If the employee has a spouse/has no spouse and more than 1 children, then all the plan need to be family plans
    - Otherwise if they do have separate insurance, need to mention the insurance company name and what kind of plans chosen will depend on that(single/couple/family)
-> While selecting the plans make sure than the plans have the same province as the employee address
-> Some plans also take age into consideration
-> A person can choose only one plan in different categories (ex. health and dental insurance, Executive benefits etc.)
-> If the plan is classic Bronze or Silver then we can include mental and health wellness plans(GroupBenefitz EAP 2.0 or Complete wellness) as well,
  If the plans are All-in Bronze/Silver/Gold then we can't include mental and health wellness plans
-> Atlantic and western provinces contain ( AB, BC, NB, NS, SK and QC)
-> Some plans are available for multiple provinces, so while filling out the excel file need to keep that in mind
-> Plan name - GroupBenefitz High Cost Drugs - there are two versions for it, one for ontario and other for rest of provinces (single/family)

Employee Age Selection:
-> Anywhere between 18 and 60

Child Graduation Date Selection:
-> From today to 3 years if they are above 15

Hiring Selection:
-> Anywhere from 10 years ago till today

Plan enrollment date Selection:
-> First of the every month from next month onwards
"""


def generate_phone_number():
    return f"+1{random.randint(100, 999):03d}{random.randint(100, 999):03d}{random.randint(1000, 9999):04d}"


def generate_child_dob():
    start_date = datetime.date(2005, 1, 1)
    end_date = datetime.date.today()
    days_between_dates = (end_date - start_date).days
    random_number_of_days = random.randrange(days_between_dates)
    random_date = start_date + datetime.timedelta(days=random_number_of_days)
    return f"{random_date}"


def check_age(dob):
    dob = datetime.datetime.strptime(dob, '%Y-%m-%d').date()
    age = datetime.date.today() - dob
    age_in_years = age.days / 365
    return age_in_years


def generate_doh(years):
    today = datetime.date.today()
    ten_years_ago = today - datetime.timedelta(days=365 * years)
    days_between_dates = (today - ten_years_ago).days
    random_number_of_days = random.randrange(days_between_dates)
    random_date = today - datetime.timedelta(days=random_number_of_days)
    return f"{random_date}"


def generate_grad_date():
    now = datetime.date.today()
    random_time = datetime.timedelta(days=random.randint(0, 3 * 365))
    return f"{now + random_time}"


def generate_dob():
    start_date = datetime.date(1960, 1, 1)
    end_date = datetime.date(2003, 1, 1)
    days_between_dates = (end_date - start_date).days
    random_number_of_days = random.randrange(days_between_dates)
    random_date = start_date + datetime.timedelta(days=random_number_of_days)
    return f"{random_date}"


def generate_ped():
    today = datetime.date.today()
    random_month = random.randint(1, 12)
    random_first_of_month = datetime.date(today.year, random_month, 1)
    return f"{random_first_of_month}"


def generate_email(firstname, lastname, companyname):
    domain_name = ["ca", "org", "edu", "com"]

    companyname = companyname.replace('Inc.', '')
    companyname = companyname.replace('Corp.', '')
    companyname = companyname.replace('Co.', '')
    companyname = companyname.replace("Corporation", '')
    companyname = companyname.replace('Incorporated', '')
    companyname = companyname.replace('Corp', '')
    companyname = companyname.replace('&', '')
    companyname = companyname.replace('Inc', '')
    companyname = companyname.replace('Company', '')
    companyname = companyname.replace('Trust', '')
    companyname = companyname.replace('Companies', '')
    companyname = companyname.replace('.', '')
    companyname = companyname.replace(',', '')
    companyname = companyname.strip()
    companyname = companyname.replace(' ', '')
    companyname = companyname.lower()

    email = f"{firstname.lower()}.{lastname.lower()}@{companyname}.{random.choice(domain_name)}"

    return email


class ReadingExcel:
    def __init__(self):
        self.data_list = []

    def createList(self, column_str, sheet):
        for column_str, sheet in zip(column_str, sheet):
            temp_list = []
            for cell in sheet[column_str]:
                if cell.value is not None:
                    temp_list.append(cell.value)
            temp_list.pop(0)
            self.data_list.append(temp_list)
        return self.data_list


class DataGenerator:
    def __init__(self):
        self.generated_data = {}

        fields = ['first_name', 'last_name', 'company_name', 'job_title', 'gender', 'date_of_hiring',
                  'hours_per_week', 'date_of_birth', 'phone_number', 'email', 'street_address_line1', 'province',
                  'city', 'postal_code', 'working_20hours', 'provincial_health_coverage', 'planEnrollmentDate',
                  'advisorName', 'paymentMethod', 'having_spouse', 'spouse_first_name',
                  'spouse_last_name', 'spouse_date_of_birth', 'spouse_gender', 'is_spouse_having_healthcard',
                  'spouse_carrier_name', 'no_of_children', 'no_of_plans', 'totalAmount']

        for i in range(1, 10):
            child_fields = [f'child{i}_first_name', f'child{i}_last_name', f'child{i}_gender', f'child{i}_date_of_birth', f'child{i}_is_child_having_healthcard', f'child{i}_child_carrier_name',
                            f'child{i}_enrolledInUniversity', f'child{i}_isDisabled', f'child{i}_graduationDay']
            fields += child_fields

        for i in range(1, 10):
            plan_fields = [f'plan{i}_planname', f'plan{i}_details', f'plan{i}_planPrice', f'plan{i}_tax', f'plan{i}_totalAmount']
            fields += plan_fields

        for field in fields:
            self.generated_data[field] = []

        self.first_names, self.last_names, self.company_names, self.job_titles, self.insurance_companies, self.street_addresses, self.plans = self.generated_lists()
        self.gender_list = ['Male', 'Female', 'Undisclosed', 'Non-Binary']
        self.gender_perc = [48, 45, 4, 3]
        self.payment_method_list = ["Credit Card", "Pre-Authorized Debit"]
        self.having_spouse_list = [True, False]
        self.having_spouse_perc = [70, 30]

    def generated_lists(self):
        workbook1 = openpyxl.load_workbook('book1.xlsx')
        sheet1 = workbook1["Sheet 1"]

        workbook2 = openpyxl.load_workbook('dummy address.xlsx')
        sheet2 = workbook2["Sheet1"]

        workbook3= openpyxl.load_workbook('plans.xlsx', data_only=True)
        sheet3 = workbook3['Sheet 1']

        excel_reader = ReadingExcel()
        columns_strs = ['A', 'B', 'C', 'D', 'E', 'A']
        sheets = [sheet1, sheet1, sheet1, sheet1, sheet1, sheet2]
        lists = excel_reader.createList(columns_strs, sheets)

        first_name_list, last_name_list, company_name_list, job_title_list, insurance_company_list, address_list_temp = lists

        address_list = []
        i = 0
        while i < len(address_list_temp) - 1:
            if i % 2 == 0:
                temp_list = f"{address_list_temp[i]}^{address_list_temp[i + 1]}"
                address_list.append(temp_list)
            i += 1

        plans_list = []
        for row in sheet3.rows:
            values = []
            for cell in row:
                if cell.value != None:
                    values.append(cell.value)
            if len(values) > 0:
                plans_list.append(values)
        plans_list.pop(0)

        return first_name_list, last_name_list, company_name_list, job_title_list, insurance_company_list, address_list, plans_list

    def generate_data(self, num_records):
        for _ in range(num_records):
            self.generated_data['first_name'].append(random.choice(self.first_names))
            self.generated_data['last_name'].append(random.choice(self.last_names))
            self.generated_data['company_name'].append(random.choice(self.company_names))
            self.generated_data['job_title'].append(random.choice(self.job_titles))
            self.generated_data['gender'].append(random.choices(self.gender_list, weights=self.gender_perc)[0])
            self.generated_data['date_of_hiring'].append(generate_doh(10))
            self.generated_data['hours_per_week'].append(random.randint(20, 60))
            self.generated_data['date_of_birth'].append(generate_dob())
            self.generated_data['phone_number'].append(generate_phone_number())
            self.generated_data['email'].append(generate_email(self.generated_data['first_name'][-1], self.generated_data['last_name'][-1], self.generated_data['company_name'][-1]))
            temp_adr = random.choice(self.street_addresses)
            temp_adr = temp_adr.split("^")
            street_address_line1 = temp_adr[0].strip()
            city = temp_adr[1].split(',')[0].strip()
            temp_pr = temp_adr[1].split(',')[1].strip()
            province = temp_pr[0:2]
            postal_code = temp_pr[2:len(temp_pr)].strip()
            self.generated_data['street_address_line1'].append(street_address_line1)
            self.generated_data['province'].append(province)
            self.generated_data['city'].append(city)
            self.generated_data['postal_code'].append(postal_code)
            self.generated_data['working_20hours'].append(True)
            self.generated_data['provincial_health_coverage'].append(True)
            self.generated_data['planEnrollmentDate'].append(generate_ped())
            self.generated_data['advisorName'].append(f"{random.choice(self.first_names)} {random.choice(self.last_names)}")
            self.generated_data['paymentMethod'].append(random.choice(self.payment_method_list))
            having_spouse_choice = random.choices(self.having_spouse_list, weights=self.having_spouse_perc)[0]
            self.generated_data['having_spouse'].append(having_spouse_choice)
            if having_spouse_choice:
                spouse_first_name = random.choice(self.first_names)
                spouse_last_name = random.choice(self.last_names)
                spouse_date_of_birth = generate_dob()
                spouse_gender = random.choices(self.gender_list, weights=self.gender_perc)[0]
                is_spouse_having_healthcard = random.choices([True, False], weights=[35, 65])[0]
                if is_spouse_having_healthcard:
                    spouse_carrier_name = random.choice(self.insurance_companies)
                else:
                    spouse_carrier_name = None
            else:
                spouse_first_name = None
                spouse_last_name = None
                spouse_date_of_birth = None
                spouse_gender = None
                is_spouse_having_healthcard = None
                spouse_carrier_name = None
            self.generated_data['spouse_first_name'].append(spouse_first_name)
            self.generated_data['spouse_last_name'].append(spouse_last_name)
            self.generated_data['spouse_date_of_birth'].append(spouse_date_of_birth)
            self.generated_data['spouse_gender'].append(spouse_gender)
            self.generated_data['is_spouse_having_healthcard'].append(is_spouse_having_healthcard)
            self.generated_data['spouse_carrier_name'].append(spouse_carrier_name)
            no_of_children = random.randint(0, 9)
            self.generated_data['no_of_children'].append(no_of_children)
            for i in range(1, 10):
                if i <= no_of_children:
                    child_first_name = random.choice(self.first_names)
                    child_last_name = random.choice(self.last_names)
                    child_gender = random.choices(self.gender_list, weights=self.gender_perc)[0]
                    child_dob = generate_child_dob()
                    is_child_having_healthcard = random.choices([True, False], weights=[35, 65])[0]
                    if is_child_having_healthcard:
                        child_carrier_name = random.choice(self.insurance_companies)
                    else:
                        child_carrier_name = None
                    child_age = check_age(child_dob)
                    is_child_disabled = random.choices([True, False], weights=[25, 75])[0]
                    if child_age >= 15:
                        child_enrolled = True
                        child_graduation_date = generate_grad_date()
                    else:
                        child_enrolled = False
                        child_graduation_date = None
                else:
                    child_first_name = None
                    child_last_name = None
                    child_gender = None
                    child_dob = None
                    is_child_having_healthcard = None
                    child_carrier_name = None
                    child_enrolled = None
                    is_child_disabled = None
                    child_graduation_date = None
                self.generated_data[f'child{i}_first_name'].append(child_first_name)
                self.generated_data[f'child{i}_last_name'].append(child_last_name)
                self.generated_data[f'child{i}_gender'].append(child_gender)
                self.generated_data[f'child{i}_date_of_birth'].append(child_dob)
                self.generated_data[f'child{i}_is_child_having_healthcard'].append(is_child_having_healthcard)
                self.generated_data[f'child{i}_child_carrier_name'].append(child_carrier_name)
                self.generated_data[f'child{i}_enrolledInUniversity'].append(child_enrolled)
                self.generated_data[f'child{i}_isDisabled'].append(is_child_disabled)
                self.generated_data[f'child{i}_graduationDay'].append(child_graduation_date)

            no_of_plans = random.randint(1, 9)
            self.generated_data['no_of_plans'].append(no_of_plans)
            total_amount = 0

            for i in range(1, 10):
                plan = random.choice(self.plans)
                if i <= no_of_plans:
                    plan_name = plan[0].strip()
                    plan_detail = plan[1].strip()
                    plan_price = plan[2]
                    plan_tax = plan[3]
                    plan_total = plan[4]
                    total_amount += plan_total
                else:
                    plan_name = None
                    plan_detail = None
                    plan_price = None
                    plan_tax = None
                    plan_total = None
                self.generated_data[f'plan{i}_planname'].append(plan_name)
                self.generated_data[f'plan{i}_details'].append(plan_detail)
                self.generated_data[f'plan{i}_planPrice'].append(plan_price)
                self.generated_data[f'plan{i}_tax'].append(plan_tax)
                self.generated_data[f'plan{i}_totalAmount'].append(plan_total)

            self.generated_data['totalAmount'].append(total_amount)


class ExcelWriter:
    def __init__(self, data_generator):
        self.data_generator = data_generator

    def fill_excel(self, generated_list):
        wb = openpyxl.load_workbook('test_data.xlsx')
        sheet = wb['Sheet1']
        for row in range(2, 1001):
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header in self.data_generator.generated_data:
                    sheet.cell(row=row, column=col).value = self.data_generator.generated_data[header][row - 2]
        wb.save('test_data.xlsx')


if __name__ == '__main__':
    data_gen = DataGenerator()
    data_gen.generate_data(1001)
    writer = ExcelWriter(data_gen)
    writer.fill_excel(data_gen.generated_data)
