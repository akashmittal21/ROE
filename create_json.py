import sys
import openpyxl
import json
import os


class Person:
    def __init__(self, row, spouse_list, child_list, plans_list):
        self.data = {}
        self.header = sheet[1]
        self.row = int(row)
        self.populate_data(spouse_list, child_list, plans_list)
        self.data['fileName'] = f"test_data{self.row}.pdf"
        self.data['filePath'] = f"{os.getcwd()}/"
        self.data['disclouseradvisor'] = True
        self.data['termsandconditions'] = True
        self.data['apt'] = ""
        self.data['signature'] = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAASwAAACWCAYAAABkW7XSAAAAAXNSR0IArs4c6QAAD0tJREFUeF7tnWfoNUcVxp9oFHuNCAasiGgIURSTWNCgCEowxo4KNlBULIl+UREVEUEklmDAIMSCKFGMwYAfxBYLioqK5ZNGg8FYIJYEG4LyyO7rONm9d+/d2Ttz9v4WXl7e/7u7c+Z3Zp//mTPtJHFBAAIQCELgpCB2YiYEIAABIVg0AghAIAwBBCuMqzAUAhBAsGgDEIBAGAIIVhhXYSgEIIBg0QYgAIEwBBCsMK7CUAhAAMGiDUAAAmEIIFhhXIWhEIAAgkUbgAAEwhBAsMK4CkMhAAEEizYAAQiEIYBghXEVhkIAAggWbQACEAhDAMEK4yoMhQAEECzaAAQgEIYAghXGVRgKAQggWLQBCEAgDAEEK4yrMBQCEECwaAMQgEAYAghWGFdhKAQggGDRBiAAgTAEEKwwrsJQCEAAwaINQAACYQggWGFchaEQgACCRRuAAATCEECwwrgKQyEAAQSLNgABCIQhgGCFcRWGQgACCBZtAAIQCEMAwQrjKgyFAAQQLNoABCAQhgCCFcZVGAoBCCBYtAEIQCAMAQQrjKswFAIQQLBoAxCAQBgCCFYYV2EoBCCAYNEGIACBMAQQrDCuwlAIQADBog1AAAJhCCBYYVyFoRCAAIJFG4AABMIQQLDCuApDIQABBIs2AAEIhCGAYIVxFYZCAAIIFm0AAhAIQwDBCuMqDIUABBAs2gAEIBCGAIIVxlUYCgEIIFi0AQhAIAwBBCuMqzAUAhBAsGgDEIBAGAIIVhhXYSgEIIBg0QYgAIEwBBCsMK7CUAhAAMGiDUAAAmEIIFhhXLWzoQ+VdJqkMyU9StKNkj4p6dKd38QDEGiEAILViCMKmmGh+qKkU0be+X1Jn5H0bUlfLVgur4LA4gQQrMURH7yA90l67cRS3ybp7RPv5TYIVCeAYFV3QVED7iLpl5L899TLUdYlkj499QHug0AtAghWLfLLlOuI6a3Zqz8q6WJJd5T0dEmvHijaonXOMibxVgiUI4BglWPZwpv+JOnOiSFfk/T4zLAndwL2gOznFixyWi14ERtGCSBY62kcFqavZNU5X9LnRqp4kaQLkv9zLssRGhcEmiWAYDXrmp0Ny7uDQ9FV+tKzJX0r+cFfsuhsZwN4AAJLE0CwliZ8uPfngjUlYvqHpFt3Jl4v6V6HM5eSILA7AQRrd2atPvE6Se/dsYv3Q0lndM9si8harTd2HREBBGs9zn6apCuS6lwpyT/bdDnJ/jgEaz2NYO01QbDW4+E86X6tpPtuqd6vJN2nu8fTH160HhzUZI0EEKx1efXfWXU2+TfPeW0aUVwXJWoTlsDaBMszvL2Wzh+uIwxHEMd0pYL1G0mnjlQ+nxE/JRo7Jo7UtVECkQTLQnReNxHSXZ2+u+PJkreXdKsBxhasj0h6vyTft/bre5IenlRyyL8Wq6slnZ7cR3S19paxkvq1LlhPkfQ8SY+ekI/Z5hJ3gdYuXJ44ms5sz2ev+/88kTSdDU/ualvL4f+bIdCiYDkCcCTlYXpHVSWvmyR9OJvhXfL9td+V56V+LumBnVFD6wz/KOn+RxJ91vYN5Rcg0JJg+bf/C7uh+LHdBn7UrXdz9y5d93YHSf7z9y4S8/Me8epHwHJUl0l6SQF+Lb4iT7w7gvJltun1TUnP73J9LdYDmyBwMwK1BcvC4g/J0dTYELyjosslvWOPJLqjCr877QL1EKbMBI/YZJyzS8XpX5JOTiry5+6XAgudI3r3yG2uJVgWJ2+DsmnejyMD51vGFu9OdZ1F0aL1Fkm3yB56cZeUn/quCPc5qvyZpNsNGOsI1cw9w50LAuEIHFqwtgmVPyhHCP5TelQvnwluZ7mM+y1QVq2G4Do6//ccSbfNjPiFpEesqK61GFNuRQKHEixHOd6215FOnp9yF6UXqaV/8zu6cP4qvdYSZX1e0rkb2pKneDwMwar4tVH0bAKHECwn0y0SeY7K0ZT3H3eXr3Q0tQlMvkh4ypq72aAXfIF/AXgNYb5R3+8k3TMr16w954oLAiEJLClY/oDeKOlJGRnvCuBkeK2krz9wD+f3lyMPdwsjXo4YvUNDGrVeI+kZXZ7KAuUuYnpFHGxwPV0PtxnPpeM6UgJLCJaF6p3dWXgp1tpCldqSLvr1z5fgsGSTMmMPWuRRVS5GFjLXNR8lvTDbimZJW+e820LleqbRuX+5HNuSqzkMV/Vs6Q91aHLilyW9vrGRqXwJS5SPwB+uI6p82xj/MnBXdygH6Mm3jkxy0Wp1D3fX0flOi9VQvnOXE4FW9bFSmbKRxXckPTKB6mOjLGAeYm/t+puk2yRGeelPul1wa/b6I32PpGdJulNinBctW6i2Tf2waP1goFJOwi890DGFpUXKUbnXQT5o4IFa+c4ptnPPAQmUirDyNWwt50n8UTjCSi8P9/tE5NausdFVC5V/GXh0der1Mkkfym72YIcjrRqi1U8adiQ1tATrkKPHUxlyX2UCJQQrF6tfS7p35XptKt45kfx0mNYEyx+wu0Xu+uVdICedbf8+I6tD0zr8Hkdah8oL9XPxhupmv5WaMNxwE8S0fQnMFaz8A2j9QM6ho7DM7u6SbtgXYqHnLFLerngs4vCHbKGaKyz5tA6b7wjLkdY+Ijil+hYp1y1f0O5I0d1ZtxvXq0akN8V+7mmEwFzB+rqkx3R18XC6G+V1jdQtN8ORyk8GNrX7hKQXVLDZAuUDICyi/jO0ltLdIn/QJYQqrWK+3tD/t8QvG9fJEeFTM74WX9tQa2pLBXdTZAkCcwXrxm6XhN6WQ+45ZQFyROC/+25T/ncfMfjvd0l67gC0pUfLbJOFyQLlD/iJ3ZYuQ2v9evOWXKLUlzEkWqVyj0NLsDxBt18bulQkV+Kb4B0NE5grWPlWJq6qG2Mf2p/SnXX3h5HRn9pobL8Z5N2s/N/91stT7PWzvZBuOwQifZ+nJvQf9Nxu3xQ7XSdHOP0xX/0zc5Yq9bm3dFH7pikXU+zkHgicIDBXsDxl4cHw3JmAt3xxFHVVJxq1ukZjorXrdIehiawI1c7Ngge2EZgrWG7wHu1xV3Bss7xtNkz5f0doPkrdOShfaXfQ/+4jEkc2Q90Nz196SFLQZyX9uOumpRFR2qVMR+d8T19mWkb/M+/qectublc+qudoM00o1xKnMc5DojV1uoN971HXdFpCSysaprQt7glEYK5g9VXtd/h0VyDvYvTdLt/rJLI/hk1dHn/gfbcy31l0H7ROqH88ebBUnmYfW1p9Zki0fivp7BFfDS2ZQaha9e6K7ColWCkSN37nbtKopxayi7L92xGrcU/Yb86h9SdB+858jpYjKi8NSnNzCFWt1n2E5S4hWC1h/FS3mZ1t8prGJ7RkXIO2WLS85U+6xbJF6xWSXp4ttkaoGnTg2k1au2B5Gc4HJf1+YC7Q2n07p36p0OfvsVD1+5jNKYNnIbAzgbUL1s5AeOAEAS+WTpPp/+xOGvJEWy4IVCGAYFXB3myh7hK+SdJLJd1twEpPNvU8LS4IVCGAYFXB3lyhYwuSvazmLElnJhZ7CosHL7ggcHACCNbBkTdV4NjOpd+V9OxupNdi5qkm6QaAr5J0SVM1wZijIIBgHYWbb1bJMaEa2xFiaANAjxpeepz4qHUtAghWLfJ1yt1VqFIrvYPsMzOzmddWx49HWyqCdRyuHzvA1jsoeI+qKYut7yHJXUEvxUmv0xrdBvs4PHtktUSw1u1wj/pZYCxK6eWF1/7ZPusa800QibLW3Yaaqh2C1ZQ7ihoztN7PazktVLvsBT9kVLqt0BIb/xUFwcvWQwDBWo8v+5pYqLy0ZujMQs9QL7F5ng+iTXeluGuh967PG9SoKAEEqyjOqi87tzv/MRcqL6WxiE3JU02pgEXPB2Sk1677Z00ph3sgcDMCCFb8RmExsoDkR2UtsTh56BAP58OGjumKT5YaNEcAwWrOJZMNGspR+eE5CfVNhVuUfKRb2hV0Tsw/LxW9Ta48Nx4nAQQrnt/HhMrnQXqvKv8peVmg/M50n/b+/edPOHW6pC2868gJIFhxGoDnUl02kEx3DeYcrrqJgNcNvlnSyQM3eQ6XN/TjgsDBCCBYB0M9qyBHN45y8v3i3f3z/y1xAOnlkrwX/tBlsXK5JUYcZ4Hh4eMigGC1729HOfnsclu91IRNi5TL8wz2/LJAepRw7jyu9qljYZMEEKwm3XLCqKHDTktN/sxr7hHAV45EVU6qX0C+qu3GcgzWIVjtetkHQpyXmWexsrCU7AJ6lO8Dkh47gsKLni+UdF27qLDsWAggWO152nmqKwaS69d2Se4SYuUyPBveeaixOVQ/lWSxYrO+9trI0VqEYLXleo8Eeq5TfsS9c0eOrOYmufuj5D26lyfwexLXSPqYpIsl3dAWHqw5dgIIVjstwAJytaTTM5NKjMiNrS9Mi/JJ2F+S9G5J17eDBUsg8D8CCFY7rWFoNNA7gA5N2JxitaM0d/u8O8NYNOVupnNlHvljtvoUqtxTlQCCVRX/icLzNXrevsWJbgvJLpeFyYl6i1y+CDp9j9cZegSS6Qm70OXe6gQQrOou+G++ymcAplGQZ67nm+5tstQ5qddIOmfDTY6mepEimqrvdyzYgwCCtQe0wo941O+M5J2OfjZFR/2tTqD3I31jXT4n693l858So4uFq87rILAbAQRrN16l73bXzesD+2vbVi2Oxvou39B0hJsk+Yguj/JZpOaOKpauL++DwCwCCNYsfLMfTqOrsa1atomUjXByvo+kZhvFCyDQKgEEq65n0r3RvfQlTbI7grpK0qkjJnq6Qy9SRFJ1/UjpByKAYB0I9EgxqWD1vnAC3TuIDuWx3GV04txCReK8ru8ovQIBBKsC9K7IdCqDhcti5OR5Psv9r5K+IOkNiFQ9Z1FyGwQQrHp+yBPuuSXOabmLWOqkm3o1pWQIFCKAYBUCucdr3PXzIuf86udLIVR7QOWRdRNAsOr616J1lqRvSPKUBCfPmS9V1yeU3jABBKth52AaBCDw/wQQLFoEBCAQhgCCFcZVGAoBCCBYtAEIQCAMAQQrjKswFAIQQLBoAxCAQBgCCFYYV2EoBCCAYNEGIACBMAQQrDCuwlAIQADBog1AAAJhCCBYYVyFoRCAAIJFG4AABMIQQLDCuApDIQABBIs2AAEIhCGAYIVxFYZCAAIIFm0AAhAIQwDBCuMqDIUABBAs2gAEIBCGAIIVxlUYCgEIIFi0AQhAIAyB/wBrag61Kfam7gAAAABJRU5ErkJggg=="

    def get_column(self, headerString):
        for cell in self.header:
            if cell.value == headerString:
                return cell.column

    def populate_data(self, spouse_list, child_list, plan_list):
        for col in range(1, 22):
            key = sheet.cell(row=1, column=col).value
            value = sheet.cell(row=self.row, column=col).value
            self.data[key] = value

        self.populate_spouse(spouse_list)
        self.populate_children(child_list)
        self.populate_plans(plan_list)
        self.populate_totals()

    def populate_spouse(self, spouse_list):
        col = self.get_column('having_spouse')
        if sheet.cell(row=self.row, column=col).value:
            spouse_details = {}
            for key in spouse_list:
                col += 1
                spouse_value = sheet.cell(row=self.row, column=col).value
                spouse_details[key] = spouse_value
            self.data['spouse_details'] = spouse_details
        else:
            self.data['spouse_details'] = {}

    def populate_children(self, child_list):
        col = self.get_column('no_of_children')
        no_of_children = sheet.cell(row=self.row, column=col).value
        if no_of_children > 0:
            i = 1
            child_array = []
            while i <= no_of_children:
                children_details = {}
                for key in child_list:
                    col += 1
                    child_value = sheet.cell(row=self.row, column=col).value
                    children_details[key] = child_value
                child_array.append(children_details)
                i += 1
            self.data['children_details'] = child_array
        else:
            self.data['children_details'] = {}

    def populate_plans(self, plans_list):
        col = self.get_column('no_of_plans')
        no_of_plans = int(sheet.cell(row=self.row, column=col).value)
        i = 1
        plan_array = []
        while i <= no_of_plans:
            plan = {}
            for key in plans_list:
                col += 1
                plan_value = sheet.cell(row=self.row, column=col).value
                plan[key] = plan_value
            plan_array.append(plan)
            i += 1
        self.data['plans'] = plan_array

    def populate_totals(self):
        col = self.get_column('totalAmount')
        self.data['totalAmount'] = sheet.cell(row=self.row, column=col).value


if __name__ == '__main__':
    # read data from spreadsheet
    wb = openpyxl.load_workbook('test_data.xlsx', data_only=True)
    sheet = wb['Sheet1']
    input_row = sys.argv[1]

    # Checking if the input is an integer
    try:
        input_row = int(input_row)
    except ValueError:
        print(f"Error: Please enter an integer value!!")

    # Checking the max rows in Excel sheet
    max_row = sheet.max_row

    # Checking if input is between the available rows that are in spreadsheet
    if input_row > max_row or input_row < 2:
        print(f"Error! Please enter the correct row number \nAvailable rows 2 - {max_row}")
    else:
        spouse_field_list = ['first_name', 'last_name', 'date_of_birth', 'gender', 'is_spouse_having_healthcard', 'spouse_carrier_name']
        child_field_list = ['first_name','last_name','gender','date_of_birth','is_child_having_healthcard','child_carrier_name','enrolledInUniversity','isDisabled','graduationDay']
        plans_field_list = ['planname','details','planPrice','tax','totalPrice']

        person = Person(input_row, spouse_field_list, child_field_list, plans_field_list)
        overall_data = person.data

        # Convert the data to a JSON string
        json_data = json.dumps(overall_data)

        # Open a file for writing
        with open('test_data.json', 'w') as f:
            # Write the JSON string to the file
            f.write(json_data)

